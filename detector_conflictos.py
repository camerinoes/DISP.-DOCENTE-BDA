"""
Detector de conflictos de horario
BDA Analytics — Universidad de Lima

Cruza la disponibilidad de docentes (extraída de Word files) contra
la propuesta de horarios (CSV de la hoja Propuesta del Excel maestro).

Uso:
    python detector_conflictos.py \
        --propuesta "PROPUESTA 2026-0 - Propuesta.csv" \
        --disponibilidad disponibilidad_consolidada.xlsx \
        --salida conflictos.xlsx

Si no se pasa --disponibilidad, solo analiza conflictos dentro de la propuesta.
"""

import csv
import argparse
import re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


# Mapeo de columnas del CSV a días de la semana
DIAS_CSV = {
    "LUNES": 16,
    "MARTES": 17,
    "MIÉRCOLES": 18,
    "JUEVES": 19,
    "VIERNES": 20,
    "SÁBADO": 21,
}

COLORES = {
    "header":    PatternFill("solid", fgColor="1F3864"),
    "rojo":      PatternFill("solid", fgColor="FFB3B3"),
    "amarillo":  PatternFill("solid", fgColor="FFE699"),
    "verde":     PatternFill("solid", fgColor="C6EFCE"),
    "gris":      PatternFill("solid", fgColor="F2F2F2"),
}
FONT_HEADER = Font(color="FFFFFF", bold=True, size=10)
CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
IZQUIERDA = Alignment(horizontal="left", vertical="center", wrap_text=True)


def hora_a_set(texto):
    """Convierte '14-18' a {14, 15, 16, 17}."""
    if not texto or texto.strip() in ("-", ""):
        return set()
    m = re.match(r"(\d+)\s*[-–]\s*(\d+)", str(texto).strip())
    if not m:
        return set()
    return set(range(int(m.group(1)), int(m.group(2))))


def leer_propuesta(ruta_csv):
    """Lee el CSV de la propuesta y devuelve lista de filas como dicts."""
    filas = []
    with open(ruta_csv, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        encabezado = next(reader)
        for row in reader:
            if len(row) < 10:
                continue
            # Columnas clave (índices del CSV de la Propuesta)
            # SIST(0) OK(1) NIV(2) COD(3) CURSO(4) CAR(5) CRED(6) SEC(7)
            # DOCENTE(8) COORD(9) DIA(10) HRS(11) -(12) TIPO(13) GPO(14)
            # JP(15) LUNES(16) MARTES(17) MIERCOLES(18) JUEVES(19) VIERNES(20) SABADO(21)
            try:
                fila = {
                    "nivel":    row[2].strip() if len(row) > 2 else "",
                    "codigo":   row[3].strip() if len(row) > 3 else "",
                    "curso":    row[4].strip() if len(row) > 4 else "",
                    "car":      row[5].strip() if len(row) > 5 else "",
                    "cred":     row[6].strip() if len(row) > 6 else "",
                    "seccion":  row[7].strip() if len(row) > 7 else "",
                    "docente":  row[8].strip() if len(row) > 8 else "",
                    "tipo":     row[13].strip() if len(row) > 13 else "",
                    "grupo":    row[14].strip() if len(row) > 14 else "",
                    "jp":       row[15].strip() if len(row) > 15 else "",
                    "lunes":    hora_a_set(row[16]) if len(row) > 16 else set(),
                    "martes":   hora_a_set(row[17]) if len(row) > 17 else set(),
                    "miercoles":hora_a_set(row[18]) if len(row) > 18 else set(),
                    "jueves":   hora_a_set(row[19]) if len(row) > 19 else set(),
                    "viernes":  hora_a_set(row[20]) if len(row) > 20 else set(),
                    "sabado":   hora_a_set(row[21]) if len(row) > 21 else set(),
                    "_raw": row,
                }
                if fila["curso"] and fila["nivel"]:
                    filas.append(fila)
            except Exception:
                continue
    return filas


def horas_docente(fila):
    """Devuelve dict {dia: set_horas} para un docente/fila."""
    return {
        "LUNES":     fila["lunes"],
        "MARTES":    fila["martes"],
        "MIÉRCOLES": fila["miercoles"],
        "JUEVES":    fila["jueves"],
        "VIERNES":   fila["viernes"],
        "SÁBADO":    fila["sabado"],
    }


def detectar_conflictos_propuesta(filas):
    """
    Detecta tres tipos de conflictos dentro de la propuesta:
    1. Mismo docente, mismo día, horas superpuestas (doble asignación)
    2. Mismo nivel, mismo día, horas superpuestas (cursos del mismo nivel que se pisan)
    3. Misma sección, mismo docente, teoría duplicada
    """
    conflictos = []

    # ── 1. Doble asignación del mismo docente ──────────────────────────────
    por_docente = defaultdict(list)
    for f in filas:
        if f["docente"] and f["tipo"] in ("TEO", ""):
            por_docente[f["docente"]].append(f)

    for docente, asignaciones in por_docente.items():
        for i in range(len(asignaciones)):
            for j in range(i + 1, len(asignaciones)):
                a, b = asignaciones[i], asignaciones[j]
                ha, hb = horas_docente(a), horas_docente(b)
                for dia in ha:
                    solapamiento = ha[dia] & hb[dia]
                    if solapamiento:
                        conflictos.append({
                            "tipo": "DOBLE ASIGNACIÓN DOCENTE",
                            "gravedad": "ALTA",
                            "docente": docente,
                            "curso_a": f"{a['curso']} (Sec {a['seccion']})",
                            "curso_b": f"{b['curso']} (Sec {b['seccion']})",
                            "nivel_a": a["nivel"],
                            "nivel_b": b["nivel"],
                            "dia": dia,
                            "horas": str(sorted(solapamiento)),
                            "detalle": f"El docente tiene dos asignaciones simultáneas el {dia} en horas {sorted(solapamiento)}",
                        })

    # ── 2. Cursos del mismo nivel que se pisan ─────────────────────────────
    por_nivel = defaultdict(list)
    for f in filas:
        if f["nivel"] and f["tipo"] in ("TEO", ""):
            por_nivel[f["nivel"]].append(f)

    for nivel, cursos_nivel in por_nivel.items():
        for i in range(len(cursos_nivel)):
            for j in range(i + 1, len(cursos_nivel)):
                a, b = cursos_nivel[i], cursos_nivel[j]
                if a["codigo"] == b["codigo"] and a["seccion"] == b["seccion"]:
                    continue  # misma sección, no es conflicto
                ha, hb = horas_docente(a), horas_docente(b)
                for dia in ha:
                    solapamiento = ha[dia] & hb[dia]
                    if solapamiento:
                        clave = tuple(sorted([
                            f"{a['curso']}|{a['seccion']}",
                            f"{b['curso']}|{b['seccion']}"
                        ])) + (dia,)
                        # Evitar duplicados
                        ya_existe = any(
                            c.get("_clave") == clave for c in conflictos
                        )
                        if not ya_existe:
                            conflictos.append({
                                "tipo": "CRUCE DE CURSOS MISMO NIVEL",
                                "gravedad": "MEDIA",
                                "docente": "",
                                "curso_a": f"{a['curso']} (Sec {a['seccion']})",
                                "curso_b": f"{b['curso']} (Sec {b['seccion']})",
                                "nivel_a": nivel,
                                "nivel_b": nivel,
                                "dia": dia,
                                "horas": str(sorted(solapamiento)),
                                "detalle": f"Nivel {nivel}: dos cursos se superponen el {dia} en horas {sorted(solapamiento)}",
                                "_clave": clave,
                            })

    return conflictos


def leer_disponibilidad_excel(ruta_xlsx):
    """Lee el Excel de disponibilidad generado por extractor_disponibilidad.py."""
    wb = openpyxl.load_workbook(ruta_xlsx, read_only=True)
    ws = wb["Disponibilidad"]
    rows = list(ws.rows)
    encabezado = [c.value for c in rows[0]]

    disponibilidad = {}
    for row in rows[1:]:
        vals = [c.value for c in row]
        if not vals or not vals[0]:
            continue
        nombre = str(vals[0]).strip().upper()
        disp = {}
        for dia in ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO"]:
            horas_disponibles = set()
            for h in ["7-8","8-9","9-10","10-11","11-12","12-13","13-14",
                      "14-15","15-16","16-17","17-18","18-19","19-20","20-21","21-22"]:
                col_name = f"{dia}\n{h}"
                if col_name in encabezado:
                    idx = encabezado.index(col_name)
                    if idx < len(vals) and vals[idx] in ("X", "x"):
                        horas_disponibles.update(hora_a_set(h))
            disp[dia] = horas_disponibles
        disponibilidad[nombre] = disp
    wb.close()
    return disponibilidad


def detectar_conflictos_disponibilidad(filas, disponibilidad_docentes):
    """
    Detecta asignaciones donde el docente está fuera de su disponibilidad declarada.
    """
    conflictos = []
    for f in filas:
        docente_raw = f["docente"].strip().upper()
        if not docente_raw:
            continue

        # Intentar match flexible (apellido)
        match_key = None
        for key in disponibilidad_docentes:
            # Comparar por primer apellido
            if docente_raw.split("/")[0] in key or key.split()[0] in docente_raw:
                match_key = key
                break

        if not match_key:
            continue

        disp = disponibilidad_docentes[match_key]
        horas_asig = horas_docente(f)

        for dia, horas in horas_asig.items():
            horas_no_disponibles = horas - disp.get(dia, set())
            if horas_no_disponibles:
                conflictos.append({
                    "tipo": "FUERA DE DISPONIBILIDAD",
                    "gravedad": "ALTA",
                    "docente": f["docente"],
                    "curso_a": f"{f['curso']} (Sec {f['seccion']})",
                    "curso_b": "",
                    "nivel_a": f["nivel"],
                    "nivel_b": "",
                    "dia": dia,
                    "horas": str(sorted(horas_no_disponibles)),
                    "detalle": f"Asignado en {dia} horas {sorted(horas_no_disponibles)} pero no marcó disponibilidad ahí",
                })
    return conflictos


def exportar_conflictos(conflictos, ruta_salida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conflictos"

    encabezado = ["TIPO", "GRAVEDAD", "DOCENTE", "CURSO A", "NIVEL A",
                  "CURSO B", "NIVEL B", "DÍA", "HORAS", "DETALLE"]
    ws.append(encabezado)
    for col_idx, titulo in enumerate(encabezado, 1):
        c = ws.cell(1, col_idx)
        c.fill = COLORES["header"]
        c.font = FONT_HEADER
        c.alignment = CENTRO

    por_gravedad = {"ALTA": [], "MEDIA": [], "BAJA": []}
    for conf in conflictos:
        por_gravedad[conf.get("gravedad", "BAJA")].append(conf)

    def escribir_grupo(lista, color):
        for conf in lista:
            ws.append([
                conf.get("tipo", ""),
                conf.get("gravedad", ""),
                conf.get("docente", ""),
                conf.get("curso_a", ""),
                conf.get("nivel_a", ""),
                conf.get("curso_b", ""),
                conf.get("nivel_b", ""),
                conf.get("dia", ""),
                conf.get("horas", ""),
                conf.get("detalle", ""),
            ])
            for col_idx in range(1, len(encabezado) + 1):
                ws.cell(ws.max_row, col_idx).fill = color
                ws.cell(ws.max_row, col_idx).font = Font(size=9)
                ws.cell(ws.max_row, col_idx).alignment = IZQUIERDA

    escribir_grupo(por_gravedad["ALTA"],   COLORES["rojo"])
    escribir_grupo(por_gravedad["MEDIA"],  COLORES["amarillo"])
    escribir_grupo(por_gravedad["BAJA"],   COLORES["verde"])

    anchos = [28, 8, 30, 35, 8, 35, 8, 12, 15, 55]
    for col_idx, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["TIPO DE CONFLICTO", "CANTIDAD"])
    tipos = defaultdict(int)
    for c in conflictos:
        tipos[c["tipo"]] += 1
    for tipo, count in sorted(tipos.items(), key=lambda x: -x[1]):
        ws2.append([tipo, count])
    ws2.append(["TOTAL", len(conflictos)])

    wb.save(ruta_salida)
    print(f"Conflictos guardados en: {ruta_salida}")


def get_column_letter(n):
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def main():
    parser = argparse.ArgumentParser(description="Detector de conflictos de horario")
    parser.add_argument("--propuesta", required=True, help="CSV de la propuesta de horarios")
    parser.add_argument("--disponibilidad", default=None, help="Excel de disponibilidad docente (opcional)")
    parser.add_argument("--salida", default="conflictos.xlsx", help="Archivo Excel de salida")
    args = parser.parse_args()

    print("Leyendo propuesta...")
    filas = leer_propuesta(args.propuesta)
    print(f"  {len(filas)} filas cargadas")

    print("Detectando conflictos en la propuesta...")
    conflictos = detectar_conflictos_propuesta(filas)

    if args.disponibilidad and Path(args.disponibilidad).exists():
        print("Cruzando con disponibilidad docente...")
        disp = leer_disponibilidad_excel(args.disponibilidad)
        conflictos += detectar_conflictos_disponibilidad(filas, disp)

    print(f"  {len(conflictos)} conflictos encontrados")
    exportar_conflictos(conflictos, args.salida)

    print("\nResumen:")
    tipos = defaultdict(int)
    for c in conflictos:
        tipos[c["tipo"]] += 1
    for tipo, count in sorted(tipos.items(), key=lambda x: -x[1]):
        print(f"  {tipo}: {count}")


if __name__ == "__main__":
    main()
